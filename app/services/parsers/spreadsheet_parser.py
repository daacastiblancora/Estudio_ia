import csv
from typing import List
from langchain_core.documents import Document
from app.services.parser_registry import BaseParser
from app.core.logging import logger

try:
    import openpyxl
except ImportError:
    openpyxl = None


class SpreadsheetParser(BaseParser):
    """Parses Excel (.xlsx) and CSV files. Each sheet/file becomes a Document."""
    supported_extensions = ["xlsx", "xls", "csv"]

    def parse(self, file_path: str, source_name: str) -> List[Document]:
        if file_path.lower().endswith(".csv"):
            return self._parse_csv(file_path, source_name)
        return self._parse_excel(file_path, source_name)

    def _parse_excel(self, file_path: str, source_name: str) -> List[Document]:
        if openpyxl is None:
            logger.error("openpyxl not installed. Run: pip install openpyxl")
            return []

        docs = []
        wb = openpyxl.load_workbook(file_path, read_only=True, data_only=True)
        total_sheets = len(wb.sheetnames)

        for idx, sheet_name in enumerate(wb.sheetnames, 1):
            ws = wb[sheet_name]
            rows = []
            for row in ws.iter_rows(values_only=True):
                cells = [str(c) if c is not None else "" for c in row]
                # Skip empty rows
                if any(c.strip() for c in cells):
                    rows.append(" | ".join(cells))

            if rows:
                text = self._inject_header(source_name, "Hoja", idx)
                text += f"Hoja: {sheet_name}\n\n"
                text += "\n".join(rows)

                docs.append(Document(
                    page_content=text,
                    metadata=self._make_metadata(
                        source=source_name, page=idx, total=total_sheets,
                        fmt="xlsx", section=sheet_name
                    ),
                ))

        wb.close()
        return docs

    def _parse_csv(self, file_path: str, source_name: str) -> List[Document]:
        try:
            with open(file_path, "r", encoding="utf-8", errors="replace") as f:
                reader = csv.reader(f)
                rows = []
                for row in reader:
                    if any(cell.strip() for cell in row):
                        rows.append(" | ".join(row))
        except Exception as e:
            logger.error(f"CSV parse error: {e}")
            return []

        if not rows:
            return []

        text = self._inject_header(source_name, "Página", 1)
        text += "\n".join(rows)

        return [Document(
            page_content=text,
            metadata=self._make_metadata(
                source=source_name, page=1, total=1, fmt="csv"
            ),
        )]
